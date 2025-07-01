from flask import Flask, request, jsonify, Blueprint, Response
from pymongo import MongoClient
import configparser
from datetime import datetime
import logging
import csv
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Initialize Blueprint for transaction routes
transaction_routes = Blueprint('transaction_routes', __name__)

# Read config file
config = configparser.ConfigParser()
config.read('config.ini')


# MongoDB connection
client = MongoClient(MONGO_URI)
wallet_db = client["cryptonel_wallet"]
user_transactions_collection = wallet_db["user_transactions"]

# API prefix
API_PREFIX = '/api'

@transaction_routes.route(f'{API_PREFIX}/admin/user/transactions', methods=['GET'])
def get_user_transactions():
    """
    Get all transactions for a specific user by their user ID
    This endpoint is accessible only by admin users
    """
    try:
        # Get user_id from request parameters
        user_id = request.args.get('user_id')
        
        if not user_id:
            return jsonify({
                "success": False,
                "message": "يجب توفير معرف المستخدم",
                "error": "missing_user_id"
            }), 400
        
        # Find user in database
        user = user_transactions_collection.find_one({"user_id": user_id})
        
        if not user:
            return jsonify({
                "success": False,
                "message": "لم يتم العثور على المستخدم",
                "error": "user_not_found"
            }), 404
        
        # Extract transactions from user document
        transactions = user.get("transactions", [])
        
        # طباعة معلومات عن عدد المعاملات
        print(f"Found {len(transactions)} transactions for user {user_id}")
        
        # تحقق من وجود المعاملات
        if not transactions:
            print(f"No transactions found for user {user_id}")
        else:
            print(f"First transaction: {transactions[0]}")
        
        # Sort transactions by timestamp (newest first)
        # استخدام datetime.fromisoformat بدلا من المقارنة المباشرة
        try:
            for tx in transactions:
                if isinstance(tx.get("timestamp"), dict) and "$date" in tx["timestamp"]:
                    # تحويل التاريخ من الصيغة الخاصة بـ MongoDB إلى datetime بايثون
                    tx["original_timestamp"] = tx["timestamp"]
                    tx["timestamp"] = tx["timestamp"]["$date"]
            
            # التصنيف بعد تحضير البيانات
            transactions.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
        except Exception as e:
            print(f"Error sorting transactions: {e}")
            # في حالة الخطأ نعيد المعاملات دون تصنيف
            pass
        
        # استخراج بيانات المستخدم من موقع آخر إذا كان ضروريًا
        user_balance = "0.00000000"
        username = "غير معروف"
        try:
            # محاولة الوصول إلى بيانات المستخدم من مجموعة users
            users_collection = wallet_db["users"]
            user_details = users_collection.find_one({"user_id": user_id})
            if user_details:
                user_balance = user_details.get("balance", "0.00000000")
                username = user_details.get("username", "غير معروف")
                print(f"Found user details with username: {username} and balance: {user_balance}")
            else:
                print(f"User details not found in users collection for user_id: {user_id}")
        except Exception as e:
            print(f"Error getting user details: {e}")
        
        # Prepare user data
        user_data = {
            "user_id": user.get("user_id"),
            "username": username,
            "public_address": user.get("public_address", "غير معروف"),
            "created_at": user.get("created_at"),
            "balance": user_balance,
            "ban": user.get("ban", False),
            "wallet_lock": user.get("wallet_lock", False),
        }
        
        # طباعة التشخيص
        print(f"Found user: {user.get('user_id')}")
        print(f"Transaction count: {len(transactions)}")
        
        # Return the transaction data
        return jsonify({
            "success": True,
            "user": user_data,
            "transactions": transactions,
            "transaction_count": len(transactions)
        })
        
    except Exception as e:
        logging.error(f"Error fetching user transactions: {str(e)}")
        return jsonify({
            "success": False,
            "message": "حدث خطأ أثناء جلب بيانات المعاملات",
            "error": str(e)
        }), 500 

@transaction_routes.route(f'{API_PREFIX}/admin/user/transactions/export', methods=['GET'])
def export_user_transactions():
    """
    تصدير معاملات المستخدم بتنسيق Excel
    """
    try:
        # الحصول على معرف المستخدم من المعلمات
        user_id = request.args.get('user_id')
        export_format = request.args.get('format', 'excel')  # excel أو csv
        
        # الحصول على المعرفات المحددة للتصدير (إذا وجدت)
        selected_ids = request.args.get('selected')
        selected_transaction_ids = selected_ids.split(',') if selected_ids else None
        
        # الحصول على ملاحظات المشرف (إذا وجدت)
        admin_notes = request.args.get('notes', '')
        
        if not user_id:
            return jsonify({
                "success": False,
                "message": "يجب توفير معرف المستخدم",
                "error": "missing_user_id"
            }), 400
        
        # البحث عن المستخدم في قاعدة البيانات
        user = user_transactions_collection.find_one({"user_id": user_id})
        
        if not user:
            return jsonify({
                "success": False,
                "message": "لم يتم العثور على المستخدم",
                "error": "user_not_found"
            }), 404
        
        # استخراج المعاملات من وثيقة المستخدم
        transactions = user.get("transactions", [])
        
        # تصفية المعاملات حسب المعرفات المحددة (إذا وجدت)
        if selected_transaction_ids:
            transactions = [tx for tx in transactions if tx.get('transaction_id') in selected_transaction_ids]
            print(f"Filtered transactions by selection, found {len(transactions)} transactions")
        
        # الحصول على بيانات المستخدم الإضافية من مجموعة المستخدمين
        users_collection = wallet_db["users"]
        user_details = users_collection.find_one({"user_id": user_id})
        username = "غير معروف"
        balance = "0.00000000"
        email = "غير متوفر"
        account_type = "غير متوفر"
        
        if user_details:
            username = user_details.get("username", "غير معروف")
            balance = user_details.get("balance", "0.00000000")
            email = user_details.get("email", "غير متوفر")
            account_type = user_details.get("account_type", "غير متوفر")
            print(f"Found user details for export with username: {username} and balance: {balance}")
        else:
            print(f"User details not found in users collection for export user_id: {user_id}")
        
        # ترتيب المعاملات حسب الطابع الزمني (الأحدث أولاً)
        for tx in transactions:
            if isinstance(tx.get("timestamp"), dict) and "$date" in tx["timestamp"]:
                tx["timestamp"] = tx["timestamp"]["$date"]
        
        transactions.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
        
        # تحويل البيانات إلى تنسيق قابل للقراءة
        formatted_transactions = []
        for tx in transactions:
            timestamp = tx.get("timestamp")
            if isinstance(timestamp, str):
                try:
                    timestamp = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                except:
                    try:
                        timestamp = datetime.strptime(timestamp, "%Y-%m-%dT%H:%M:%S.%f%z")
                    except:
                        timestamp = "تاريخ غير صحيح"
            
            status_mapping = {
                "completed": "مكتملة",
                "pending": "قيد التنفيذ",
                "canceled": "ملغية"
            }
            
            type_mapping = {
                "sent": "إرسال",
                "received": "استلام"
            }
            
            formatted_transactions.append({
                "رقم المعاملة": tx.get("transaction_id", ""),
                "النوع": type_mapping.get(tx.get("type", ""), tx.get("type", "")),
                "المبلغ": tx.get("amount", "0"),
                "الرسوم": tx.get("fee", "0"),
                "التاريخ والوقت": timestamp if isinstance(timestamp, datetime) else timestamp,
                "الحالة": status_mapping.get(tx.get("status", ""), tx.get("status", "")),
                "معرف الطرف الآخر": tx.get("counterparty_id", ""),
                "عنوان الطرف الآخر": tx.get("counterparty_address", "")
            })
        
        # بناء ملف إكسل
        wb = Workbook()
        ws = wb.active
        ws.title = "معاملات المستخدم"
        
        # إعداد اتجاه الصفحة للغة العربية
        ws.sheet_view.rightToLeft = True
        
        # أنماط التنسيق
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # حدود الخلايا
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # معلومات المستخدم في الأعلى - تحسين التنسيق
        ws.merge_cells('A1:H1')
        ws['A1'] = f"تقرير معاملات المستخدم: {username} ({user_id})"
        ws['A1'].font = Font(name='Arial', size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        
        # القسم الأول: معلومات المستخدم
        info_fill = PatternFill(start_color="E8F1FB", end_color="E8F1FB", fill_type="solid")
        label_font = Font(name='Arial', size=11, bold=True)
        value_font = Font(name='Arial', size=11)
        
        # الصف الثاني: معرف المستخدم
        ws.merge_cells('A2:B2')
        ws['A2'] = "معرف المستخدم:"
        ws['A2'].font = label_font
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws['A2'].fill = info_fill
        ws['A2'].border = thin_border
        
        ws.merge_cells('C2:D2')
        ws['C2'] = user_id
        ws['C2'].font = value_font
        ws['C2'].alignment = Alignment(horizontal='right', vertical='center')
        ws['C2'].fill = info_fill
        ws['C2'].border = thin_border
        
        # الرصيد
        ws.merge_cells('E2:F2')
        ws['E2'] = "الرصيد الحالي:"
        ws['E2'].font = label_font
        ws['E2'].alignment = Alignment(horizontal='left', vertical='center')
        ws['E2'].fill = info_fill
        ws['E2'].border = thin_border
        
        ws.merge_cells('G2:H2')
        ws['G2'] = f"{balance} CRN"
        ws['G2'].font = value_font
        ws['G2'].alignment = Alignment(horizontal='right', vertical='center')
        ws['G2'].fill = info_fill
        ws['G2'].border = thin_border
        
        # الصف الثالث: اسم المستخدم
        ws.merge_cells('A3:B3')
        ws['A3'] = "اسم المستخدم:"
        ws['A3'].font = label_font
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
        ws['A3'].fill = info_fill
        ws['A3'].border = thin_border
        
        ws.merge_cells('C3:D3')
        ws['C3'] = username
        ws['C3'].font = value_font
        ws['C3'].alignment = Alignment(horizontal='right', vertical='center')
        ws['C3'].fill = info_fill
        ws['C3'].border = thin_border
        
        # البريد الإلكتروني
        ws.merge_cells('E3:F3')
        ws['E3'] = "البريد الإلكتروني:"
        ws['E3'].font = label_font
        ws['E3'].alignment = Alignment(horizontal='left', vertical='center')
        ws['E3'].fill = info_fill
        ws['E3'].border = thin_border
        
        ws.merge_cells('G3:H3')
        ws['G3'] = email
        ws['G3'].font = value_font
        ws['G3'].alignment = Alignment(horizontal='right', vertical='center')
        ws['G3'].fill = info_fill
        ws['G3'].border = thin_border
        
        # إضافة تاريخ التقرير
        ws.merge_cells('A4:H4')
        report_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = f"تم إنشاء التقرير في: {report_date}"
        ws['A4'].font = Font(name='Arial', size=10, italic=True)
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A4'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        ws['A4'].border = thin_border
        
        # إضافة ملاحظات المشرف إذا وجدت
        current_row = 5
        if admin_notes.strip():
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = f"ملاحظات المشرف: {admin_notes}"
            ws[f'A{current_row}'].font = Font(name='Arial', size=11)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'A{current_row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            ws[f'A{current_row}'].border = thin_border
            current_row += 1
        
        # إضافة معلومات عن المعاملات المحددة
        if selected_transaction_ids:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = f"تم تصدير {len(transactions)} معاملة من أصل {len(user.get('transactions', []))} معاملة"
            ws[f'A{current_row}'].font = Font(name='Arial', size=10)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
            ws[f'A{current_row}'].border = thin_border
            current_row += 1
        
        # إضافة مساحة
        ws.row_dimensions[current_row].height = 10
        current_row += 1
        
        # رؤوس الأعمدة - تحسين التنسيق
        headers = ["رقم المعاملة", "النوع", "المبلغ", "الرسوم", "التاريخ والوقت", "الحالة", "معرف الطرف الآخر", "عنوان الطرف الآخر"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # ضبط عرض الأعمدة
        column_widths = [30, 12, 15, 15, 25, 15, 30, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # إضافة البيانات
        for row_num, tx in enumerate(formatted_transactions, current_row + 1):
            # نمط الصف
            row_fill = PatternFill(start_color="F5F9FD", end_color="F5F9FD", fill_type="solid") if row_num % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # تلوين حسب نوع المعاملة
            type_fill = None
            type_font = Font(name='Arial', size=11, bold=True)
            
            if tx["النوع"] == "إرسال":
                type_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                type_font = Font(name='Arial', size=11, bold=True, color="C00000")
            elif tx["النوع"] == "استلام":
                type_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                type_font = Font(name='Arial', size=11, bold=True, color="006100")
            
            # تلوين حسب الحالة
            status_fill = None
            status_font = Font(name='Arial', size=11)
            
            if tx["الحالة"] == "مكتملة":
                status_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                status_font = Font(name='Arial', size=11, color="006100", bold=True)
            elif tx["الحالة"] == "قيد التنفيذ":
                status_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                status_font = Font(name='Arial', size=11, color="9C5700", bold=True)
            elif tx["الحالة"] == "ملغية":
                status_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
                status_font = Font(name='Arial', size=11, color="C00000", bold=True)
            
            # إضافة بيانات الصف
            for col_num, (key, value) in enumerate(tx.items(), 1):
                cell = ws.cell(row=row_num, column=col_num)
                
                # تنسيق التاريخ
                if key == "التاريخ والوقت" and isinstance(value, datetime):
                    cell.value = value.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    cell.value = value
                
                # تطبيق التنسيق الأساسي
                cell.border = thin_border
                
                if col_num != 2 and col_num != 6:  # ليس النوع أو الحالة
                    cell.fill = row_fill
                    
                # تنسيق المبلغ
                if key == "المبلغ":
                    if tx["النوع"] == "إرسال":
                        cell.font = Font(name='Arial', size=11, color="C00000", bold=True)
                    else:
                        cell.font = Font(name='Arial', size=11, color="006100", bold=True)
                
                # تنسيقات خاصة حسب العمود
                if key == "النوع" and type_fill:
                    cell.fill = type_fill
                    cell.font = type_font
                elif key == "الحالة" and status_fill:
                    cell.fill = status_fill
                    cell.font = status_font
                
                # محاذاة النص
                if key in ["رقم المعاملة", "معرف الطرف الآخر", "عنوان الطرف الآخر"]:
                    cell.alignment = Alignment(horizontal='right')
                elif key in ["المبلغ", "الرسوم"]:
                    cell.alignment = Alignment(horizontal='center')
                    # إضافة وحدة العملة
                    if key == "المبلغ":
                        cell.value = f"{value} CRN"
                    if key == "الرسوم" and float(value) > 0:
                        cell.value = f"{value} CRN"
                else:
                    cell.alignment = Alignment(horizontal='center')
        
        # مصادقة المستند
        ws.protection.sheet = True
        ws.protection.password = "cryptonel"
        
        # حفظ الملف مؤقتًا
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # تجهيز ملف الإكسل للتنزيل
        filename = f"transactions_{username}_{user_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
        
    except Exception as e:
        logging.error(f"Error exporting user transactions: {str(e)}")
        return jsonify({
            "success": False,
            "message": "حدث خطأ أثناء تصدير بيانات المعاملات",
            "error": str(e)
        }), 500 

@transaction_routes.route(f'{API_PREFIX}/admin/user/toggle-ban', methods=['POST'])
def toggle_user_ban():
    """
    تغيير حالة حظر المستخدم
    هذه الواجهة متاحة فقط للمشرفين
    """
    try:
        # الحصول على البيانات من الطلب
        data = request.get_json()
        user_id = data.get('user_id')
        new_ban_status = data.get('ban')
        
        if user_id is None or new_ban_status is None:
            return jsonify({
                "success": False,
                "message": "يجب توفير معرف المستخدم وحالة الحظر الجديدة",
                "error": "missing_parameters"
            }), 400
        
        # البحث عن المستخدم في مجموعة المستخدمين
        users_collection = wallet_db["users"]
        user = users_collection.find_one({"user_id": user_id})
        
        if not user:
            return jsonify({
                "success": False,
                "message": "لم يتم العثور على المستخدم",
                "error": "user_not_found"
            }), 404
        
        # تحديث حالة الحظر للمستخدم
        users_collection.update_one(
            {"user_id": user_id},
            {"$set": {"ban": new_ban_status}}
        )
        
        # تحديث حالة الحظر في مجموعة معاملات المستخدم أيضًا
        user_transactions_collection.update_one(
            {"user_id": user_id},
            {"$set": {"ban": new_ban_status}}
        )
        
        # سجل الإجراء
        print(f"User {user_id} ban status changed to: {new_ban_status}")
        
        return jsonify({
            "success": True,
            "message": "تم تحديث حالة حظر المستخدم بنجاح",
            "new_ban_status": new_ban_status
        })
        
    except Exception as e:
        logging.error(f"Error toggling user ban: {str(e)}")
        return jsonify({
            "success": False,
            "message": "حدث خطأ أثناء تحديث حالة حظر المستخدم",
            "error": str(e)
        }), 500

@transaction_routes.route(f'{API_PREFIX}/admin/user/toggle-wallet-lock', methods=['POST'])
def toggle_wallet_lock():
    """
    تغيير حالة قفل محفظة المستخدم
    هذه الواجهة متاحة فقط للمشرفين
    """
    try:
        # الحصول على البيانات من الطلب
        data = request.get_json()
        user_id = data.get('user_id')
        new_lock_status = data.get('wallet_lock')
        
        if user_id is None or new_lock_status is None:
            return jsonify({
                "success": False,
                "message": "يجب توفير معرف المستخدم وحالة قفل المحفظة الجديدة",
                "error": "missing_parameters"
            }), 400
        
        # البحث عن المستخدم في مجموعة المستخدمين
        users_collection = wallet_db["users"]
        user = users_collection.find_one({"user_id": user_id})
        
        if not user:
            return jsonify({
                "success": False,
                "message": "لم يتم العثور على المستخدم",
                "error": "user_not_found"
            }), 404
        
        # تحديث حالة قفل المحفظة للمستخدم
        users_collection.update_one(
            {"user_id": user_id},
            {"$set": {"wallet_lock": new_lock_status}}
        )
        
        # تحديث حالة قفل المحفظة في مجموعة معاملات المستخدم أيضًا
        user_transactions_collection.update_one(
            {"user_id": user_id},
            {"$set": {"wallet_lock": new_lock_status}}
        )
        
        # سجل الإجراء
        print(f"User {user_id} wallet lock status changed to: {new_lock_status}")
        
        return jsonify({
            "success": True,
            "message": "تم تحديث حالة قفل محفظة المستخدم بنجاح",
            "new_lock_status": new_lock_status
        })
        
    except Exception as e:
        logging.error(f"Error toggling wallet lock: {str(e)}")
        return jsonify({
            "success": False,
            "message": "حدث خطأ أثناء تحديث حالة قفل محفظة المستخدم",
            "error": str(e)
        }), 500 
